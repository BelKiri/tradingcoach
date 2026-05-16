"""
Universal Excel trade history parser.

Handles ANY .xlsx/.xls trade history export from any broker or platform.
Auto-detects header rows, column mapping, date formats, and skips non-trade rows.
Returns standardized trade dicts matching the `trades` table schema.
"""

from __future__ import annotations

import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl

from tradecoach.parsers._format_spec import (
    ALIASES,
    BUY_TYPES,
    DATE_FORMATS,
    EXCEL_NATIVE_DATETIME_SENTINEL,
    PRICE_NAMES,
    REQUIRED_CANONICAL_FIELDS,
    SELL_TYPES,
    SKIP_TYPES,
    TIME_HEADER_DUPLICATE_LABEL,
    has_mapped_time_column,
    header_detection_keywords,
    normalize_header_label,
    primary_date_sample_column,
)


class XlsxParseError(Exception):
    """Raised when an Excel file cannot be parsed as trade history."""


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_xlsx(
    content: bytes,
) -> list[dict[str, Any]]:
    """Parse an Excel trade history file.

    Args:
        content: Raw file bytes (.xlsx or .xls).

    Returns:
        List of trade dicts matching the `trades` table schema.

    Raises:
        XlsxParseError: If the file cannot be parsed.
    """
    try:
        import io
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise XlsxParseError(f"Cannot open Excel file: {exc}")

    ws = wb.active
    if ws is None or ws.max_row is None or ws.max_row < 2:
        raise XlsxParseError("Workbook has no data")

    # Read all rows as lists of values
    all_rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    if not all_rows:
        raise XlsxParseError("No rows found")

    # Find header row
    header_idx = _find_header_row(all_rows)
    if header_idx is None:
        raise XlsxParseError("Cannot find header row with trade columns")

    header = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]
    col_map = _map_columns(header)
    _validate_columns(col_map)

    data_rows = all_rows[header_idx + 1:]
    date_fmt = _detect_date_format(data_rows, col_map)

    trades: list[dict[str, Any]] = []
    errors: list[str] = []

    for i, row in enumerate(data_rows, start=header_idx + 2):
        # Pad row if shorter than header
        if len(row) < len(header):
            row.extend([None] * (len(header) - len(row)))

        try:
            trade = _parse_row(row, col_map, date_fmt)
        except _SkipRow:
            continue
        except Exception as exc:
            errors.append(f"Row {i}: {exc}")
            continue

        trades.append(trade)

    if not trades and errors:
        raise XlsxParseError(
            f"No valid trades parsed. Errors:\n" + "\n".join(errors[:10])
        )

    return trades


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

class _SkipRow(Exception):
    pass


def _find_header_row(rows: list[list[Any]]) -> int | None:
    """Find the row index that contains column headers."""
    all_keywords = header_detection_keywords()

    best_idx = None
    best_score = 0

    for i, row in enumerate(rows[:20]):  # only check first 20 rows
        cells = [str(c).strip().lower() for c in row if c is not None]
        score = sum(1 for c in cells if c in all_keywords)
        if score > best_score:
            best_score = score
            best_idx = i

    if best_score >= 3:
        return best_idx
    return None


def _map_columns(header: list[str]) -> dict[str, int]:
    """Map canonical field names to column indices (same rules as CSV parser)."""
    col_map: dict[str, int] = {}
    normalized = [normalize_header_label(raw) for raw in header]
    price_syn = frozenset(PRICE_NAMES)

    for i, cell in enumerate(normalized):
        if cell in price_syn:
            continue
        if cell in ALIASES:
            canonical = ALIASES[cell]
            if canonical not in col_map:
                col_map[canonical] = i

    price_indices: list[int] = []
    for i, cell in enumerate(normalized):
        if cell in PRICE_NAMES and i not in col_map.values():
            price_indices.append(i)

    if price_indices:
        if "open_price" not in col_map:
            col_map["open_price"] = price_indices[0]
        if len(price_indices) >= 2 and "close_price" not in col_map:
            col_map["close_price"] = price_indices[1]

    time_indices: list[int] = []
    for i, cell in enumerate(normalized):
        if cell == TIME_HEADER_DUPLICATE_LABEL and i not in col_map.values():
            time_indices.append(i)
    if time_indices:
        if "open_time" not in col_map:
            col_map["open_time"] = time_indices[0]
        if len(time_indices) >= 2 and "close_time" not in col_map:
            col_map["close_time"] = time_indices[1]

    return col_map


def _validate_columns(col_map: dict[str, int]) -> None:
    """Ensure minimum required columns are present."""
    missing = REQUIRED_CANONICAL_FIELDS - set(col_map.keys())
    if missing:
        raise XlsxParseError(
            f"Missing required columns: {', '.join(sorted(missing))}. "
            f"Found: {', '.join(sorted(col_map.keys()))}"
        )
    if not has_mapped_time_column(col_map):
        raise XlsxParseError(
            "Missing a recognizable date/time column. Expected at least one "
            "header such as Open Time, Close Time, Open, or Close, or two "
            "columns named Time."
        )


def _detect_date_format(
    data_rows: list[list[Any]], col_map: dict[str, int],
) -> str | None:
    """Try to detect date format from sample rows."""
    date_col = primary_date_sample_column(col_map)
    if date_col is None:
        return None

    for row in data_rows[:10]:
        if date_col >= len(row) or row[date_col] is None:
            continue
        val = row[date_col]
        # If openpyxl already parsed it as datetime, no format needed
        if isinstance(val, datetime):
            return EXCEL_NATIVE_DATETIME_SENTINEL
        val_str = str(val).strip()
        if not val_str:
            continue
        for fmt in DATE_FORMATS:
            try:
                datetime.strptime(val_str, fmt)
                return fmt
            except ValueError:
                continue

    return None


def _get_cell(row: list[Any], col_map: dict[str, int], field: str) -> Any:
    """Get cell value by canonical field name."""
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_number(value: Any) -> float | None:
    """Parse a cell value to float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s in ("—", "-", "N/A", ""):
        return None
    # Remove spaces, currency symbols
    s = s.replace(" ", "").replace("$", "").replace("€", "")
    # Handle comma as decimal separator
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    elif "," in s and "." in s:
        s = s.replace(",", "")
    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        return None


def _strip_tz(dt: datetime) -> datetime:
    """Convert to UTC and strip tzinfo to ensure naive UTC datetimes."""
    if dt.tzinfo is not None:
        from datetime import timezone
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def _parse_datetime_value(value: Any, fmt: str | None) -> str | None:
    """Parse a datetime value from Excel cell. Returns naive UTC ISO string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return _strip_tz(value).isoformat()
    s = str(value).strip()
    if not s:
        return None

    if fmt and fmt != EXCEL_NATIVE_DATETIME_SENTINEL:
        try:
            return datetime.strptime(s, fmt).isoformat()
        except ValueError:
            pass

    # Fallback: try all formats
    for f in DATE_FORMATS:
        try:
            return datetime.strptime(s, f).isoformat()
        except ValueError:
            continue
    return None


def _clean_symbol(raw: str) -> str:
    """Normalize symbol: strip suffixes like .cash, .ecn, trailing m."""
    s = raw.strip().upper()
    s = re.sub(r"[._](CASH|ECN|PRO|SB|STD|RAW|MICRO|MINI|STP|C)$", "", s, flags=re.I)
    if len(s) > 3 and s.endswith("M") and s[-2].isalpha():
        s = s[:-1]
    s = s.lstrip("#")
    return s


def _calc_pips(
    symbol: str, direction: str,
    open_price: float | None, close_price: float | None,
) -> float | None:
    """Calculate profit in pips from open/close prices."""
    if open_price is None or close_price is None:
        return None
    diff = close_price - open_price
    if direction == "sell":
        diff = -diff
    jpy = "JPY" in symbol.upper()
    if jpy:
        pips = diff * 100
    else:
        pips = diff * 10_000
    return round(pips, 1)


def _parse_row(
    row: list[Any], col_map: dict[str, int], date_fmt: str | None,
) -> dict[str, Any]:
    """Parse a single row into a standardized trade dict."""
    # Get type/direction
    type_val = _get_cell(row, col_map, "type")
    if type_val is None:
        raise _SkipRow()
    type_str = str(type_val).strip().lower()

    if not type_str or any(skip in type_str for skip in SKIP_TYPES):
        raise _SkipRow()

    if type_str in BUY_TYPES or "buy" in type_str:
        direction = "buy"
    elif type_str in SELL_TYPES or "sell" in type_str:
        direction = "sell"
    else:
        raise _SkipRow()

    # Symbol
    raw_symbol = _get_cell(row, col_map, "symbol")
    if raw_symbol is None or str(raw_symbol).strip() == "":
        raise _SkipRow()  # skip rows without symbol (summary rows)
    symbol = _clean_symbol(str(raw_symbol))

    # Lot
    lot = _parse_number(_get_cell(row, col_map, "lot"))
    if lot is None or lot <= 0:
        raise ValueError(f"Invalid lot: {_get_cell(row, col_map, 'lot')}")

    # Prices
    open_price = _parse_number(_get_cell(row, col_map, "open_price"))
    close_price = _parse_number(_get_cell(row, col_map, "close_price"))
    stop_loss = _parse_number(_get_cell(row, col_map, "stop_loss"))
    take_profit = _parse_number(_get_cell(row, col_map, "take_profit"))

    # Zero SL/TP means not set
    if stop_loss is not None and stop_loss == 0:
        stop_loss = None
    if take_profit is not None and take_profit == 0:
        take_profit = None

    # Financials
    profit = _parse_number(_get_cell(row, col_map, "profit"))
    commission = _parse_number(_get_cell(row, col_map, "commission"))
    swap = _parse_number(_get_cell(row, col_map, "swap"))

    # Pips from file (if available)
    pips_from_file = _parse_number(_get_cell(row, col_map, "pips"))

    # Dates
    opened_at = _parse_datetime_value(_get_cell(row, col_map, "open_time"), date_fmt)
    closed_at = _parse_datetime_value(_get_cell(row, col_map, "close_time"), date_fmt)

    # Ticket
    ticket_val = _get_cell(row, col_map, "ticket")
    ticket = None
    if ticket_val is not None:
        ticket_str = str(ticket_val).strip()
        # Some brokers use composite IDs like "127939260120734416" — take first part
        if ticket_str.isdigit():
            ticket = int(ticket_str)

    # Pips: prefer from file, fallback to calculated
    profit_pips = pips_from_file
    if profit_pips is None:
        profit_pips = _calc_pips(symbol, direction, open_price, close_price)

    return {
        "ticket": ticket,
        "symbol": symbol,
        "direction": direction,
        "lot": lot,
        "open_price": open_price,
        "close_price": close_price,
        "stop_loss": stop_loss,
        "take_profit": take_profit,
        "profit_pips": profit_pips,
        "profit_money": profit,
        "commission": commission,
        "swap": swap,
        "opened_at": opened_at,
        "closed_at": closed_at,
        "source": "csv",
    }
